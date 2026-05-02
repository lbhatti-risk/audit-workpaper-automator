from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import ITGCControl

# ── Colour palette ──────────────────────────────────────────────────────────
NAVY      = "1F3864"   # section header background
SKY       = "BDD7EE"   # table column header background
YELLOW    = "FFEB9C"   # gap / deficiency row highlight
RED_BG    = "FFC7CE"   # severe deficiency
GREEN_BG  = "C6EFCE"   # evidence aligned
WHITE     = "FFFFFF"
LIGHT_GREY = "F2F2F2"
DARK_TEXT  = "1F1F1F"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=WHITE, size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _thin_border() -> Border:
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _wrap(horizontal="left", vertical="top") -> Alignment:
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)


def _write_cell(ws, row: int, col: int, value, bold=False, bg=None, font_color=DARK_TEXT,
                font_size=10, wrap=True, horizontal="left", merge_to_col: int = None):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    cell.font = Font(bold=bold, color=font_color, size=font_size, name="Calibri")
    cell.alignment = Alignment(wrap_text=wrap, horizontal=horizontal, vertical="top")
    cell.border = _thin_border()
    if bg:
        cell.fill = _fill(bg)
    if merge_to_col:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=merge_to_col
        )
    return cell


def _section_header(ws, row: int, title: str, num_cols: int = 6):
    _write_cell(ws, row, 1, title, bold=True, bg=NAVY, font_color=WHITE,
                font_size=11, merge_to_col=num_cols)
    ws.row_dimensions[row].height = 20


def _col_header_row(ws, row: int, headers: list[str], bg=SKY):
    for col, h in enumerate(headers, start=1):
        _write_cell(ws, row, col, h, bold=True, bg=bg, font_color=DARK_TEXT, font_size=10)
    ws.row_dimensions[row].height = 18


# ── Cover / Summary sheet ───────────────────────────────────────────────────

def _build_cover(wb: Workbook, controls: List[ITGCControl], generated_on: str):
    ws = wb.active
    ws.title = "Cover"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "ITGC Design & Implementation – Audit Workpaper"
    title_cell.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    title_cell.fill = _fill(NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    row = 3
    info_fields = [
        ("Client", controls[0].client_name if controls else ""),
        ("Application", controls[0].application if controls else ""),
        ("Audit Period", controls[0].audit_period if controls else ""),
        ("Date Generated", generated_on),
    ]
    for label, value in info_fields:
        _write_cell(ws, row, 1, label, bold=True, bg=LIGHT_GREY)
        _write_cell(ws, row, 2, value)
        row += 1

    row += 1
    _section_header(ws, row, "CONTROLS REVIEWED", num_cols=4)
    row += 1
    _col_header_row(ws, row, ["#", "Control", "Workpaper Sheet", "Deficiencies (CD/W)"])
    row += 1

    for i, ctrl in enumerate(controls, start=1):
        bg = YELLOW if ctrl.deficiencies else WHITE
        _write_cell(ws, row, 1, i, bg=bg)
        _write_cell(ws, row, 2, ctrl.name, bg=bg)
        _write_cell(ws, row, 3, ctrl.abbrev, bg=bg)
        cd_count = len(ctrl.deficiencies)
        _write_cell(ws, row, 4, str(cd_count) if cd_count else "None identified", bg=bg)
        row += 1

    row += 1
    total_cds = sum(len(c.deficiencies) for c in controls)
    _write_cell(ws, row, 1, "Total Control Deficiencies", bold=True)
    _write_cell(ws, row, 2, str(total_cds), bold=True, bg=YELLOW if total_cds else WHITE)


# ── Individual control sheet ─────────────────────────────────────────────────

def _build_control_sheet(wb: Workbook, ctrl: ITGCControl, cd_offset: int) -> int:
    ws = wb.create_sheet(title=ctrl.abbrev)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 32

    row = 1
    # Control header bar
    ws.merge_cells(f"A{row}:F{row}")
    h = ws[f"A{row}"]
    h.value = f"{ctrl.name}  |  {ctrl.application}  |  {ctrl.client_name}  |  {ctrl.audit_period}"
    h.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    h.fill = _fill(NAVY)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 2

    # ── Section A: Tailored Procedures ──────────────────────────────────────
    _section_header(ws, row, "SECTION A – TAILORED PROCEDURES", num_cols=6)
    row += 1
    _col_header_row(ws, row, ["#", "Procedure", "Procedure Response", "Evidence Reference", "Gap?", ""])
    row += 1

    for pr in ctrl.procedure_results:
        bg = YELLOW if pr.has_gap else WHITE
        _write_cell(ws, row, 1, pr.number, bg=bg)
        _write_cell(ws, row, 2, pr.procedure_text, bg=bg)
        _write_cell(ws, row, 3, pr.response, bg=bg)
        _write_cell(ws, row, 4, pr.evidence_references, bg=bg)
        _write_cell(ws, row, 5, "Yes" if pr.has_gap else "No", bg=bg,
                    horizontal="center")
        _write_cell(ws, row, 6, "", bg=bg)
        ws.row_dimensions[row].height = 60
        row += 1

    row += 1

    # ── Section B: Process Description ──────────────────────────────────────
    _section_header(ws, row, "SECTION B – PROCESS DESCRIPTION", num_cols=6)
    row += 1
    _write_cell(ws, row, 1, ctrl.process_description, merge_to_col=6)
    ws.row_dimensions[row].height = max(60, len(ctrl.process_description) // 6)
    row += 2

    # ── Section C: Evidence Analysis ────────────────────────────────────────
    _section_header(ws, row, "SECTION C – EVIDENCE ANALYSIS", num_cols=6)
    row += 1
    _col_header_row(ws, row, ["#", "Evidence File", "Description", "Key Findings",
                               "Aligns with\nProcess?", "Concerns"])
    row += 1

    if ctrl.evidence_items:
        for ev in sorted(ctrl.evidence_items, key=lambda e: e.order):
            align_bg = {
                "Yes": GREEN_BG,
                "No": RED_BG,
                "Partial": YELLOW,
            }.get(ev.aligns_with_process, WHITE)

            _write_cell(ws, row, 1, ev.order)
            _write_cell(ws, row, 2, ev.filename)
            _write_cell(ws, row, 3, ev.description)
            _write_cell(ws, row, 4, "\n".join(f"• {f}" for f in ev.key_findings))
            _write_cell(ws, row, 5, ev.aligns_with_process, bg=align_bg, horizontal="center")
            _write_cell(ws, row, 6, ev.concerns)
            ws.row_dimensions[row].height = max(60, len(ev.description) // 5)
            row += 1
    else:
        _write_cell(ws, row, 1, "No evidence uploaded.", merge_to_col=6)
        row += 1

    row += 1

    # ── Section D: Control Deficiencies ─────────────────────────────────────
    _section_header(ws, row, "SECTION D – CONTROL DEFICIENCIES", num_cols=6)
    row += 1

    if ctrl.deficiencies:
        _col_header_row(ws, row, ["CD/W Ref", "Procedure Ref", "Description",
                                   "Severity", "Management Response", ""])
        row += 1
        for deficiency in ctrl.deficiencies:
            ref = f"CD/W-{cd_offset + deficiency.cd_number}"
            sev_bg = RED_BG if "Material" in deficiency.severity else YELLOW
            _write_cell(ws, row, 1, ref, bold=True, bg=sev_bg)
            _write_cell(ws, row, 2, f"Procedure {deficiency.procedure_number}", bg=sev_bg)
            _write_cell(ws, row, 3, deficiency.description, bg=sev_bg)
            _write_cell(ws, row, 4, deficiency.severity, bg=sev_bg)
            _write_cell(ws, row, 5, deficiency.management_response, bg=sev_bg)
            _write_cell(ws, row, 6, "", bg=sev_bg)
            ws.row_dimensions[row].height = max(60, len(deficiency.description) // 5)
            row += 1
    else:
        _write_cell(ws, row, 1, "No control deficiencies identified.", merge_to_col=6, bg=GREEN_BG)
        row += 1

    row += 1

    # ── Section E: Conclusion ────────────────────────────────────────────────
    _section_header(ws, row, "SECTION E – CONCLUSION", num_cols=6)
    row += 1
    conclusion_bg = RED_BG if ctrl.deficiencies else GREEN_BG
    _write_cell(ws, row, 1, ctrl.conclusion or "Conclusion pending.", merge_to_col=6, bg=conclusion_bg)
    ws.row_dimensions[row].height = max(60, len(ctrl.conclusion or "") // 6)

    return len(ctrl.deficiencies)


# ── Power BI flat-table export ───────────────────────────────────────────────

def _pbi_table(ws, headers: list[str], rows: list[list], table_name: str):
    """Write a flat Excel Table that Power BI can import directly."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=10, name="Calibri")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    if rows:
        last_col = get_column_letter(len(headers))
        last_row = len(rows) + 1
        tbl = Table(displayName=table_name, ref=f"A1:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 70)


def _build_pbi_export(wb: Workbook, controls: List[ITGCControl]):
    """
    Three hidden-ish sheets that Power BI can connect to for dashboards.

    PBI_Controls    — one row per control (risk heatmap source)
    PBI_Deficiencies — one row per CD/W (deficiency tracker)
    PBI_Evidence    — one row per evidence item (evidence coverage view)
    """
    # ── PBI_Controls ────────────────────────────────────────────────────────
    ws_ctrl = wb.create_sheet("PBI_Controls")
    ctrl_headers = [
        "Client", "Application", "Audit_Period", "Control_Name", "Control_Abbrev",
        "D_I_Result", "Evidence_Count", "CDW_Count", "Conclusion",
    ]
    ctrl_rows = []
    cd_offset = 0
    for ctrl in controls:
        cd_count = len(ctrl.deficiencies)
        ctrl_rows.append([
            ctrl.client_name,
            ctrl.application,
            ctrl.audit_period,
            ctrl.name,
            ctrl.abbrev,
            "Ineffective" if cd_count else "Effective",
            len(ctrl.evidence_items),
            cd_count,
            ctrl.conclusion,
        ])
        cd_offset += cd_count
    _pbi_table(ws_ctrl, ctrl_headers, ctrl_rows, "tbl_Controls")

    # ── PBI_Deficiencies ────────────────────────────────────────────────────
    ws_def = wb.create_sheet("PBI_Deficiencies")
    def_headers = [
        "Client", "Application", "Audit_Period", "Control_Name", "Control_Abbrev",
        "CDW_Ref", "Procedure_Number", "Description", "Severity", "Management_Response",
    ]
    def_rows = []
    running_cd = 0
    for ctrl in controls:
        for d in ctrl.deficiencies:
            running_cd += 1
            def_rows.append([
                ctrl.client_name,
                ctrl.application,
                ctrl.audit_period,
                ctrl.name,
                ctrl.abbrev,
                f"CD/W-{running_cd}",
                d.procedure_number,
                d.description,
                d.severity,
                d.management_response,
            ])
    if not def_rows:
        def_rows = [["No deficiencies identified"] + [""] * (len(def_headers) - 1)]
    _pbi_table(ws_def, def_headers, def_rows, "tbl_Deficiencies")

    # ── PBI_Evidence ────────────────────────────────────────────────────────
    ws_ev = wb.create_sheet("PBI_Evidence")
    ev_headers = [
        "Client", "Application", "Audit_Period", "Control_Name", "Control_Abbrev",
        "Evidence_Number", "Filename", "File_Type", "Description",
        "Aligns_With_Process", "Concerns",
    ]
    ev_rows = []
    for ctrl in controls:
        for ev in sorted(ctrl.evidence_items, key=lambda e: e.order):
            ev_rows.append([
                ctrl.client_name,
                ctrl.application,
                ctrl.audit_period,
                ctrl.name,
                ctrl.abbrev,
                ev.order,
                ev.filename,
                ev.file_type,
                ev.description,
                ev.aligns_with_process,
                ev.concerns,
            ])
    if not ev_rows:
        ev_rows = [["No evidence recorded"] + [""] * (len(ev_headers) - 1)]
    _pbi_table(ws_ev, ev_headers, ev_rows, "tbl_Evidence")


# ── Public entry point ───────────────────────────────────────────────────────

def generate_workpaper(controls: List[ITGCControl], output_path: str):
    wb = Workbook()
    generated_on = date.today().strftime("%d %B %Y")

    _build_cover(wb, controls, generated_on)

    cd_offset = 0
    for ctrl in controls:
        count = _build_control_sheet(wb, ctrl, cd_offset)
        cd_offset += count

    _build_pbi_export(wb, controls)

    wb.save(output_path)
